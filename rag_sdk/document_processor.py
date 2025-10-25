import io
import os
from typing import List, Dict, Any
from pathlib import Path
import chardet
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
import markdown
from bs4 import BeautifulSoup
from .config import settings


class DocumentProcessor:
    
    @staticmethod
    def detect_encoding(file_content: bytes) -> str:
        result = chardet.detect(file_content)
        return result['encoding'] or 'utf-8'
    
    @staticmethod
    async def extract_text_from_file(file_path: str, filename: str) -> str:
        extension = Path(filename).suffix.lower()
        
        if extension == '.pdf':
            return await DocumentProcessor._extract_from_pdf(file_path)
        elif extension in ['.docx', '.doc']:
            return await DocumentProcessor._extract_from_docx(file_path)
        elif extension in ['.xlsx', '.xls']:
            return await DocumentProcessor._extract_from_excel(file_path)
        elif extension in ['.md', '.markdown']:
            return await DocumentProcessor._extract_from_markdown(file_path)
        elif extension in ['.html', '.htm']:
            return await DocumentProcessor._extract_from_html(file_path)
        elif extension in ['.txt', '.csv', '.json', '.xml', '.log']:
            return await DocumentProcessor._extract_from_text(file_path)
        else:
            return await DocumentProcessor._extract_from_text(file_path)
    
    @staticmethod
    async def _extract_from_pdf(file_path: str) -> str:
        text_parts = []
        with open(file_path, 'rb') as file:
            pdf_reader = PdfReader(file)
            for page in pdf_reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
        return '\n\n'.join(text_parts)
    
    @staticmethod
    async def _extract_from_docx(file_path: str) -> str:
        doc = Document(file_path)
        text_parts = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_parts.append(paragraph.text)
        return '\n\n'.join(text_parts)
    
    @staticmethod
    async def _extract_from_excel(file_path: str) -> str:
        workbook = load_workbook(file_path, read_only=True)
        text_parts = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                if row_text.strip():
                    text_parts.append(row_text)
                    
        workbook.close()
        return '\n'.join(text_parts)
    
    @staticmethod
    async def _extract_from_markdown(file_path: str) -> str:
        with open(file_path, 'rb') as file:
            content = file.read()
            encoding = DocumentProcessor.detect_encoding(content)
            
        with open(file_path, 'r', encoding=encoding) as file:
            md_content = file.read()
            html = markdown.markdown(md_content)
            soup = BeautifulSoup(html, 'html.parser')
            return soup.get_text()
    
    @staticmethod
    async def _extract_from_html(file_path: str) -> str:
        with open(file_path, 'rb') as file:
            content = file.read()
            encoding = DocumentProcessor.detect_encoding(content)
            
        with open(file_path, 'r', encoding=encoding) as file:
            html_content = file.read()
            soup = BeautifulSoup(html_content, 'html.parser')
            return soup.get_text()
    
    @staticmethod
    async def _extract_from_text(file_path: str) -> str:
        with open(file_path, 'rb') as file:
            content = file.read()
            encoding = DocumentProcessor.detect_encoding(content)
            
        with open(file_path, 'r', encoding=encoding, errors='ignore') as file:
            return file.read()
    
    @staticmethod
    def split_text_into_chunks(text: str, chunk_size: int = None, chunk_overlap: int = None) -> List[str]:
        chunk_size = chunk_size or settings.chunk_size
        chunk_overlap = chunk_overlap or settings.chunk_overlap
        
        chunks = []
        start = 0
        text_length = len(text)
        
        while start < text_length:
            end = start + chunk_size
            
            if end < text_length:
                last_period = text.rfind('.', start, end)
                last_newline = text.rfind('\n', start, end)
                last_space = text.rfind(' ', start, end)
                
                boundary = max(last_period, last_newline, last_space)
                if boundary > start:
                    end = boundary + 1
            
            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            
            start = end - chunk_overlap
            if start <= 0:
                start = end
                
        return chunks
    
    @staticmethod
    def prepare_chunks_for_storage(chunks: List[str], embeddings: List[List[float]]) -> List[Dict[str, Any]]:
        prepared_chunks = []
        
        for idx, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
            prepared_chunks.append({
                'content': chunk,
                'embedding': embedding,
                'chunk_index': idx,
                'metadata': {
                    'length': len(chunk)
                }
            })
            
        return prepared_chunks

